import openpyxl
import datetime
import csv
import itertools 

SEND_TO_LIST = ["send_to_this_email@example.com", "also_send_here@example.com"]
FULL_PATH = "" # Path of the folder where you have this code, e.g. /home/[username]/EximLogStatsFolder
EMAIL_FROM = "'example'<noreply@example.com>"

today = datetime.datetime.now()
last_day_previous_month = str(today - datetime.timedelta(days=today.day))[0:10]
data_filepath = FULL_PATH + "/email_stats/" + "email_stats_" + last_day_previous_month[5:7] + "-" + last_day_previous_month[0:4] + ".csv"

dest_filepath = data_filepath.replace("csv", "xlsx")

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Summary"
ws2 = wb.create_sheet(title="Data")

emails = []
data=[]

with open(data_filepath, newline="\n") as csvfile:
    reader  = csv.DictReader(csvfile)

    for row in reader:
        if reader.line_num == 2:
            ws2.append(list(row.keys()))
        if row["Email"] not in emails:
            emails.append(row["Email"])
        row_num = reader.line_num
        i = 1;
        for key, val in row.items():
            if i < 3:
                _ = ws2.cell(column=i, row=row_num, value=val)
            if i >= 3:            
                _ = ws2.cell(column=i, row=row_num, value=int(val))
            i += 1


ws1_headers = ["Email", "Average Emails Sent Per Day", "Average Emails Received Per Day"]
ws1.append(ws1_headers)
rows = range(2, len(emails) + 2)
for (row, email) in zip(rows,emails):
    row = [email, "=ROUND(SUMIF(Data!$B$2:$B$1000,A" + str(row) + ",Data!$C$2:$C$1000) / COUNTIF(Data!$B$2:$B$1000, A" + str(row) +"), 0)", "=ROUND(SUMIF('Data'!$B$2:$B$1000,A" + str(row) + ",'Data'!$d$2:$d$1000) / COUNTIF(Data!$B$2:$B$1000, A" + str(row) +"), 0)"]
    ws1.append(row)

wb.save(filename = dest_filepath)

import smtplib
from email.message import EmailMessage

#filename = "email_stats_" + today[5:7] + "-" + today[0:4] + ".xlsx"
filename = "email_stats_" + last_day_previous_month[5:7] + "-" + last_day_previous_month[0:4] + ".xlsx"

for email_to in SEND_TO_LIST:
    msg = EmailMessage()
    msg['Subject'] = "Email Stats"
    msg['From'] = EMAIL_FROM
    msg['To'] = email_to
    msg.set_content("Email Stats Attached")

    with open(dest_filepath, 'rb') as fp:
        attachment_data = fp.read()
    msg.add_attachment(attachment_data, maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename = filename)
    
    with smtplib.SMTP('localhost') as s:
        s.send_message(msg)
